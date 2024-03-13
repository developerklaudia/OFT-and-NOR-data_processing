import openpyxl


def update_excel(file_path):
    # Load the Excel workbook
    wb = openpyxl.load_workbook(file_path)
    # Select the active sheet
    sheet = wb.active

    # Define mappings of values to be replaced
    replacements = {
        '1T': '1',
        '1Cy': '1',
        '2T': '2',
        '2Cy': '2',
        '2C': '2'
    }

    # Iterate through all cells in the sheet
    for row in sheet.iter_rows():
        for cell in row:
            # Check if cell value is in the replacements dictionary
            if cell.value in replacements:
                # Replace the cell value with the corresponding value from the replacements dictionary
                cell.value = replacements[cell.value]

    # Save the changes
    wb.save(file_path)


# Example usage:

def count_zone_cells_column(file_path, column_index, zone):
    # Load the Excel workbook
    wb = openpyxl.load_workbook(file_path)
    total_count = 0

    # Iterate through all sheets in the workbook
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # Iterate through all cells in the specified column
        for row in ws.iter_rows(min_col=column_index, max_col=column_index):
            for cell in row:
                # Check if the cell contains the specified zone
                if cell.value and zone.lower() in str(cell.value).lower():
                    total_count += 1

    return total_count


def count_zone_changes(file_path, column_index):
    # Load the Excel workbook
    wb = openpyxl.load_workbook(file_path)
    # Select the active sheet
    sheet = wb.active

    # Initialize variables
    previous_zone = None
    zone_changes = {}

    # Iterate through rows in the specified column
    for row in sheet.iter_rows(min_row=2, min_col=column_index, max_col=column_index, values_only=True):
        zone = row[0].strip().lower()  # Assuming the zones are strings

        # Check if the zone has changed
        if zone != previous_zone:
            # If the previous zone is not None (i.e., not the first row), update zone_changes
            if previous_zone is not None:
                change = (previous_zone, zone)
                if change not in zone_changes:
                    zone_changes[change] = 1
                else:
                    zone_changes[change] += 1
            previous_zone = zone

    return zone_changes


file_path = "D:\Kla\_NOR res xlsx\T19. F3, M5, NOR (Set 2, A&B)results.xlsx"
update_excel(file_path)

old_zones = ('1T', '1Cy', '2T', '2Cy', '2C')
new_zone = ('peripheral_zone')

zone1 = '2'
zone2 = '1'
zone3 = 'Nothing'

column_index = 2
changes_count = count_zone_changes(file_path, column_index)

two_t_zone_count = count_zone_cells_column(file_path, column_index, '2')
one_t_zone_count = count_zone_cells_column(file_path, column_index, '1')
Nothing_zone_count = count_zone_cells_column(file_path, column_index, 'Nothing')


result = find_first_zone(file_path, column_index, zone1, zone2)

#print(f"Number of '2' frames: {two_t_zone_count}")
print("Number of second that spent in 2 zone:", (two_t_zone_count / 30))

#print(f"Number of '1' frames: {one_t_zone_count}")
print("Number of second that spent in 1 zone:", (one_t_zone_count / 30))

#print(f"Number of 'Nothing' frames: {Nothing_zone_count}")
#print("Number of second that spent in Nothing zone:", (Nothing_zone_count / 30))
print("Zone changes:")
for change, count in changes_count.items():
    print(f"From '{change[0]}' to '{change[1]}': {count} times")

print("Total number of frames:", two_t_zone_count + one_t_zone_count + Nothing_zone_count)
print("Total minutes:", ((one_t_zone_count / 30) + (Nothing_zone_count / 30) + (two_t_zone_count / 30)) / 60)
print("Number of zone changes:", changes_count)
#print(result)
