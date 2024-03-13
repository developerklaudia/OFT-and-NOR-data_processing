import openpyxl


def replace_zone_in_excel_column(file_path, column_index, old_zones, new_zone):
    # Load the Excel workbook
    wb = openpyxl.load_workbook(file_path)

    # Iterate through all sheets in the workbook
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # Iterate through all cells in the specified column
        for row in ws.iter_rows(min_col=column_index, max_col=column_index):
            for cell in row:
                # Check if the cell contains any of the old_zones
                if cell.value:
                    for old_zone in old_zones:
                        if old_zone.lower() in str(cell.value).lower():
                            # Replace the old_zone with the new_zone
                            cell.value = str(cell.value).lower().replace(old_zone.lower(), new_zone.lower())

    # Save the modified workbook
    wb.save(file_path)


def delete_cells_with_nothing_column(file_path, column_index):
    # Load the Excel workbook
    wb = openpyxl.load_workbook(file_path)

    # Iterate through all sheets in the workbook
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # Iterate through all rows in reverse order
        for row in reversed(list(ws.iter_rows(min_col=column_index, max_col=column_index))):
            for cell in row:
                # Check if the cell contains "Nothing"
                if cell.value and "Nothing" in str(cell.value):
                    # If "Nothing" is found, delete the entire row
                    ws.delete_rows(cell.row)
                    break  # No need to continue checking other cells in this row

    # Save the modified workbook
    wb.save(file_path)


def count_zone_cells_column(file_path, column_index, zone):
    # Load the Excel workbook
    wb = openpyxl.load_workbook(file_path)
    total_count = 0

    # Iterate through all sheets in the workbook
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # Iterate through all cells in the specified column
        for row in ws.iter_rows(min_col=column_index, max_col=column_index):
            for cell in row:
                # Check if the cell contains the specified zone
                if cell.value and zone.lower() in str(cell.value).lower():
                    total_count += 1

    return total_count


def count_zone_changes1(file_path, column_index):
    # Load the Excel workbook
    wb = openpyxl.load_workbook(file_path)
    # Select the active sheet
    sheet = wb.active

    # Initialize variables to keep track of zone changes and the current zone
    zone_changes = 0
    current_zone = None

    # Iterate through rows in the specified column
    for row in sheet.iter_rows(min_row=2, min_col=column_index, max_col=column_index, values_only=True):
        zone = row[0].strip().lower()  # Assuming the zones are strings

        # Check if the zone has changed
        if zone != current_zone:
            # If the current zone is not None (i.e., not the first row), increment zone_changes
            if current_zone is not None:
                zone_changes += 1
            current_zone = zone  # Update the current zone

    # Increment zone_changes for the last zone change if the last row is not the same as the second last
    if current_zone is not None and zone != current_zone:
        zone_changes += 1

    return zone_changes


def count_zone_changes(file_path, column_index):
    # Load the Excel workbook
    wb = openpyxl.load_workbook(file_path)
    # Select the active sheet
    sheet = wb.active

    # Initialize variables to keep track of zone changes and the current zone
    zone_changes = 0
    current_zone = None

    # Iterate through rows in the specified column
    for row in sheet.iter_rows(min_row=2, min_col=column_index, max_col=column_index, values_only=True):
        zone = row[0].strip().lower()  # Assuming the zones are strings

        # Check if the zone has changed
        if zone != current_zone:
            # If the current zone is not None (i.e., not the first row), increment zone_changes
            if current_zone is not None:
                zone_changes += 1
            current_zone = zone  # Update the current zone

    # Increment zone_changes for the last zone change if the last row is not the same as the second last
    if current_zone is not None and zone != current_zone:
        zone_changes += 1

    return zone_changes


def find_first_zone(file_path, column_index, zone1, zone2):
    # Load the Excel workbook
    wb = openpyxl.load_workbook(file_path)

    first_zone1_row = None
    first_zone2_row = None

    # Iterate through all sheets in the workbook
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # Iterate through all cells in the specified column
        for row_number, row in enumerate(ws.iter_rows(min_col=column_index, max_col=column_index), start=1):
            for cell in row:
                # Check if the cell contains zone1
                if cell.value and zone1.lower() in str(cell.value).lower():
                    if first_zone1_row is None:
                        first_zone1_row = row_number
                # Check if the cell contains zone2
                elif cell.value and zone2.lower() in str(cell.value).lower():
                    if first_zone2_row is None:
                        first_zone2_row = row_number

    # Determine which zone appears first
    if first_zone1_row is not None and first_zone2_row is not None:
        if first_zone1_row < first_zone2_row:
            return f"{zone1} appears first in column {column_index}."
        else:
            return f"{zone2} appears first in column {column_index}."
    elif first_zone1_row is not None:
        return f"{zone1} appears first in column {column_index}."
    elif first_zone2_row is not None:
        return f"{zone2} appears first in column {column_index}."
    else:
        return f"Neither {zone1} nor {zone2} found in column {column_index}."


def count_zone_changes(file_path, column_index):
    # Load the Excel workbook
    wb = openpyxl.load_workbook(file_path)
    # Select the active sheet
    sheet = wb.active

    # Initialize variables
    previous_zone = None
    zone_changes = {}

    # Iterate through rows in the specified column
    for row in sheet.iter_rows(min_row=2, min_col=column_index, max_col=column_index, values_only=True):
        zone = row[0].strip().lower()  # Assuming the zones are strings

        # Check if the zone has changed
        if zone != previous_zone:
            # If the previous zone is not None (i.e., not the first row), update zone_changes
            if previous_zone is not None:
                change = (previous_zone, zone)
                if change not in zone_changes:
                    zone_changes[change] = 1
                else:
                    zone_changes[change] += 1
            previous_zone = zone

    return zone_changes

file_path = "D:\Kla\OFT res xlsx edited and with results\T5. F3, M5, OFT results.xlsx"
old_zones = ('left_zone', 'top_zone', 'right_zone', 'bottom_zone', 'right zone')
new_zone = 'peripheral_zone'
zone1 = 'peripheral_zone'
zone2 = 'central_zone'


column_index = 5  # Index of the column containing zones (1 for column A, 2 for column B, etc.)
changes_count1 = count_zone_changes1(file_path, column_index)
changes_count = count_zone_changes(file_path, column_index)
replace_zone_in_excel_column(file_path, column_index, old_zones, new_zone)
delete_cells_with_nothing_column(file_path, column_index)

peripheral_zone_count = count_zone_cells_column(file_path, column_index, 'peripheral_zone')
central_zone_count = count_zone_cells_column(file_path, column_index, 'centrsl_zone')
result = find_first_zone(file_path, column_index, zone1, zone2)

print(f"Number of 'peripheral_zone' frames: {peripheral_zone_count}")
print("Number of second that spent in peripheral zone:", (peripheral_zone_count/30))
print(f"Number of 'central_zone' frames: {central_zone_count}")
print("Number of second that spent in central zone:", (central_zone_count/30))
print("Total number of frames:", central_zone_count + peripheral_zone_count)
print("Total minutes:", ((central_zone_count/30) + (peripheral_zone_count/30))/60)
print("Number of zone changes:", changes_count)
print(result)
print("Number of zone changes in the 5th column:", changes_count1)
print("Zone changes:")
for change, count in changes_count.items():
    print(f"From '{change[0]}' to '{change[1]}': {count} times")
